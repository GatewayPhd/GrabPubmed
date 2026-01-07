import os
print("working directory:", os.getcwd())
from Bio import Entrez, Medline
import pandas as pd
import requests
import openpyxl
import time
from tqdm import trange
from bs4 import BeautifulSoup
import re

class pubmed_utils():
    def __init__(self):
        pass
        
        
    def get_main_info_into_excel(self, api_key, search_key_words, release_date_cutoff=None, paper_type="Article", grab_total=None, save_path="./paper_info.xlsx"):
        '''
        grab info from pubmed using NCBI eUtils API, save it into a excel
        支持逻辑符号: AND, OR, NOT 等
        
        Parameters:
        -----------
        api_key : str
            NCBI eUtils API key
        search_key_words : str
            Search query (支持逻辑符号，如 "wnt5a AND cancer")
        release_date_cutoff : int, optional
            发布时间范围（天数），默认为None（所有数据）
        paper_type : str, optional
            论文类型，默认为"Article"
        grab_total : int, optional
            获取论文数量，默认为None（获取所有）
        save_path : str
            Excel保存路径
        '''
        
        grab_step = 10
        
        # 构建搜索词
        search_term = search_key_words
        if paper_type:
            search_term += f" AND \"{paper_type}\"[PT]"
        
        # 步骤1: ESearch - 搜索论文
        esearch_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
        esearch_params = {
            "db": "pubmed",
            "term": search_term,
            "api_key": api_key,
            "usehistory": "y",
            "retmax": 0  # 只获取总数
        }
        
        # 添加日期范围限制
        if release_date_cutoff:
            esearch_params["reldate"] = release_date_cutoff
        
        print("Searching PubMed...")
        esearch_response = requests.get(esearch_url, params=esearch_params)
        esearch_data = esearch_response.text
        
        # 解析搜索结果
        import xml.etree.ElementTree as ET
        root = ET.fromstring(esearch_data)
        total = int(root.find("Count").text)
        webenv = root.find("WebEnv").text
        query_key = root.find("QueryKey").text
        
        print(f"Find total: {total}")
        
        if grab_total is None or grab_total > total:
            grab_total = total
        
        # 初始化Excel
        self.excel_property_dic = {token:index for index, token in enumerate(["PMID", "TI", "TA", "IF", "Quartile", "JCR_Quartile", "Top", "OA", "LR", "AB", "LID"], start=1)}
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=self.excel_property_dic["PMID"]).value = "PMID"
        ws.cell(row=1, column=self.excel_property_dic["TI"]).value = "Title"
        ws.cell(row=1, column=self.excel_property_dic["TA"]).value = "Journal"
        ws.cell(row=1, column=self.excel_property_dic["IF"]).value = "IF"
        ws.cell(row=1, column=self.excel_property_dic["Quartile"]).value = "JCR_Quartile"
        ws.cell(row=1, column=self.excel_property_dic["JCR_Quartile"]).value = "CSA_Quartile"
        ws.cell(row=1, column=self.excel_property_dic["Top"]).value = "Top"
        ws.cell(row=1, column=self.excel_property_dic["OA"]).value = "Open Access"
        ws.cell(row=1, column=self.excel_property_dic["LR"]).value = "publish_date"
        ws.cell(row=1, column=self.excel_property_dic["AB"]).value = "Abstract"
        ws.cell(row=1, column=self.excel_property_dic["LID"]).value = "DOI"

        # 步骤2: EFetch - 获取详细信息
        cur_row = 2
        efetch_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
        
        for step in trange(0, (grab_total + grab_step - 1) // grab_step, desc="getting pubmed info"):
            efetch_params = {
                "db": "pubmed",
                "retstart": step * grab_step,
                "retmax": grab_step,
                "webenv": webenv,
                "query_key": query_key,
                "rettype": "medline",
                "retmode": "text",
                "api_key": api_key
            }
            
            efetch_response = requests.get(efetch_url, params=efetch_params)
            response_text = efetch_response.text
            
            # 修复：使用正则表达式按照 PMID 行来分割记录
            # PMID行格式为: "PMID- 12345678"
            record_texts = re.split(r'\n(?=PMID- )', response_text)
            
            for record_text in record_texts:
                if not record_text.strip() or not record_text.startswith('PMID-'):
                    continue
                
                try:
                    # 解析单条记录
                    records = list(Medline.parse(record_text.split('\n')))
                    
                    for record in records:
                        if 'PMID' not in record:
                            continue
                            
                        # 写入Excel - 每个字段
                        for key in self.excel_property_dic.keys():
                            if key not in record:
                                continue
                            
                            key_info = record[key]
                            
                            # 处理列表类型的字段
                            if isinstance(key_info, list):
                                if key == 'LID':  # DOI 字段
                                    # 找到包含 [doi] 的项
                                    doi_items = [item for item in key_info if '[doi]' in item.lower()]
                                    if doi_items:
                                        key_info = doi_items[0]
                                    elif key_info:
                                        key_info = key_info[0]
                                    else:
                                        key_info = ''
                                elif key == 'LR':  # 日期字段 - 取最新的（第一个）
                                    key_info = key_info[0] if key_info else ''
                                else:
                                    # 其他列表字段不应该出现在这些关键字段中
                                    # 如果出现，用分号连接
                                    key_info = '; '.join(str(x) for x in key_info)
                            
                            ws.cell(row=cur_row, column=self.excel_property_dic[key]).value = key_info
                        
                        cur_row += 1
                        
                except Exception as e:
                    print(f"解析记录时出错: {e}")
                    continue
            
            time.sleep(0.5)  # 遵守API限制

        wb.save(save_path)
        print(f"Data saved to {save_path}")
        print(f"Total records written: {cur_row - 2}")
        
        
    def embed_IF_into_excel(self, excel_path, jcr_csa_path="E:\\Python\\GrabPubmed\\JCR_CSA_2025.xlsx"):
        '''
        从本地JCR_CSA_2025.xlsx获取IF、JCR分区、CSA分区信息并保存到excel
        使用pandas进行数据匹配，支持MedAbbr字段匹配
        '''

        # 加载JCR_CSA数据（指定字段类型防止自动转换）
        dtype_spec = {
            'CAS_Quartile': 'string', 
            'JIF_Quartile': 'string', 
            'JIF_2024': 'string', 
            'ISSN': 'string', 
            'eISSN': 'string'
        }
        jcr_csa_df = pd.read_excel(jcr_csa_path, dtype=dtype_spec)
        
        # 创建MedAbbr到信息的映射字典（转大写用于匹配）
        jcr_csa_dict = {}
        for _, row in jcr_csa_df.iterrows():
            med_abbr = row.get('MedAbbr')
            if pd.notna(med_abbr):
                key = str(med_abbr).strip().upper()
                jcr_csa_dict[key] = {
                    'JIF_2024': row.get('JIF_2024', 'N/A'),
                    'JIF_Quartile': row.get('JIF_Quartile', 'N/A'),
                    'CAS_Quartile': row.get('CAS_Quartile', 'N/A')
                }
        
        # 加载目标Excel
        query_df = pd.read_excel(excel_path)
        
        # 删除旧的IF相关列（如果存在）
        cols_to_drop = ['IF', 'JCR_Quartile', 'CSA_Quartile', 'Top', 'Open Access']
        existing_cols = [col for col in cols_to_drop if col in query_df.columns]
        if existing_cols:
            query_df = query_df.drop(columns=existing_cols)
        
        # 匹配统计
        match_stats = {
            'exact_match': 0,
            'no_match': 0
        }
        
        # 新增列用于存储匹配结果
        jif_2024_list = []
        jif_quartile_list = []
        cas_quartile_list = []
        
        # 遍历每条记录进行匹配
        for _, row in query_df.iterrows():
            journal_name = row.get('Journal')
            
            if pd.isna(journal_name):
                jif_2024_list.append('N/A')
                jif_quartile_list.append('N/A')
                cas_quartile_list.append('N/A')
                continue
            
            # 转大写匹配
            journal_key = str(journal_name).strip().upper()
            
            if journal_key in jcr_csa_dict:
                # 精确匹配
                match_info = jcr_csa_dict[journal_key]
                jif_2024_list.append(match_info['JIF_2024'])
                jif_quartile_list.append(match_info['JIF_Quartile'])
                cas_quartile_list.append(match_info['CAS_Quartile'])
                match_stats['exact_match'] += 1
            else:
                # 未匹配
                jif_2024_list.append('N/A')
                jif_quartile_list.append('N/A')
                cas_quartile_list.append('N/A')
                match_stats['no_match'] += 1
        
        # 添加新列到DataFrame
        query_df['IF'] = jif_2024_list
        query_df['JCR_Quartile'] = jif_quartile_list
        query_df['CSA_Quartile'] = cas_quartile_list
        
        # 统一缺失值表示
        query_df = query_df.fillna('N/A')
        
        # 保存结果
        query_df.to_excel(excel_path, index=False, sheet_name='Sheet')
        
        # 打印匹配报告
        total_journals = len(query_df)
        print("\n" + "="*60)
        print("期刊信息匹配报告")
        print("="*60)
        print(f"总期刊数: {total_journals}\n")
        print(f"精确匹配: {match_stats['exact_match']} ({match_stats['exact_match']/total_journals*100:.1f}%)")
        print(f"未匹配: {match_stats['no_match']} ({match_stats['no_match']/total_journals*100:.1f}%)")
        print("="*60)
        
        return query_df
    
def download_pdf(self, excel_path, pdf_savepath, IF_cutoff):
        '''
        try to download paper which IF higher than cutoff
        warning: very low successful rate
        '''
        
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Sheet"]
        base_url = "https://sci-hub.tw/"
        success_count = 0
        for cur_row in trange(2, ws.max_row+1, desc="downloading pdf"):
            IF = ws.cell(row=cur_row, column=self.excel_property_dic["IF"]).value
            pmid = ws.cell(row=cur_row, column=self.excel_property_dic["PMID"]).value
            title = ws.cell(row=cur_row, column=self.excel_property_dic["TI"]).value
            if IF=="Unknow" or float(IF)<IF_cutoff:
                continue

            file_name = pdf_savepath+pmid+"_"+title+".pdf"
            try:
                doi = ws.cell(row=cur_row, column=self.excel_property_dic["LID"]).value.split(" ")[0]
                url = base_url + doi
                getpage = requests.get(url, verify=True)
                getpage_soup = BeautifulSoup(getpage.text, "html.parser")
                src = getpage_soup.find("iframe", src=True).get_attribute_list("src")[0]
                response = requests.get("https:"+src, verify=True)
                f = open(file_name, "wb+")
                f.write(response.content)
                f.close()
                success_count += 1
            except:
                pass
            time.sleep(1)
        print("successful download: {}".format(success_count))
